
from selenium import webdriver;
import openpyxl as excel

#ChromeDriverのパスを引数に指定しChromeを起動
driver = webdriver.Chrome("C:/Users/原敬太/Documents/python_test/chromedriver")
#指定したURLに遷移
driver.get("https://www.google.co.jp")
#ドライバーを閉じる
driver.quit()

wbname = "test.xlsx"
msg = "Hello"
wb = excel.Workbook()

ws = wb.active

ws["A1"] = msg

wbname = "test.xlsx"

msg = "unko"

wb = excel.load_workbook(wbname)
ws = wb.active

ws["A3"] = msg

wb.save("test.xlsx")



