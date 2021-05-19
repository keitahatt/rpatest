import MySQLdb
import openpyxl as excel



# データベースへの接続とカーソルの生成
connection = MySQLdb.connect(
    host='localhost',
    user='root',
    passwd='tkvip24week',
    db='pytest',
# テーブル内部で日本語を扱うために追加
    charset='utf8'
)


cursor = connection.cursor()

#データベースのデータを抽出
cursor.execute("SELECT * FROM rpa_data")

#配列宣言
titles = []
ul = []
niti = []

tmp_num = 1

#抽出したデータを配列に追加 (test[0]はDBのタイトル、test[1]はURLでDBの左から0.1.2....となっていく)
for test in cursor:
    titles.append(test[0])
    ul.append(test[1])
    niti.append(test[2])
    print(test)
    tmp_num += 1

print(tmp_num)
print(titles)
print(ul)
print(niti)


#エクセル生成
wbname = "sports.xlsx"
wb = excel.Workbook()
ws = wb.active

#エクセルシート情報
ws["A1"].value = 'タイトル'
ws["B1"].value = 'URL'
ws["C1"].value = '日時'


wb.create_sheet()

for i in range(1, tmp_num):
    ws.cell(column=1, row=i+1, value=titles[i-1]) #配列titlesの要素をA列に出力
    ws.cell(column=2, row=i+1, value=ul[i-1]) #配列ulの要素をB列に出力
    ws.cell(column=3, row=i+1, value=niti[i-1]) #配列nitiの要素をC列に出力

wb.save("sports.xlsx")


# 接続を閉じる
connection.close()