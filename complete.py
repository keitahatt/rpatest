import MySQLdb
import openpyxl as excel
import datetime
from bs4.element import Comment
import requests
import glob
from bs4 import BeautifulSoup
import re
from selenium import webdriver


# データベースへの接続とカーソルの生成
connection = MySQLdb.connect(
    host='localhost',
    user='root',
    passwd='tkvip24week',
    db='pytest',
# テーブル内部で日本語を扱うために追加
    charset='utf8'
)


cursor = connection.cursor()

driver = webdriver.Chrome("C:/Users/原敬太/Documents/python_test/chromedriver")
url = 'https://news.yahoo.co.jp/categories/sports/'
driver.get(url)
#ドライバーを閉じる
driver.quit()

response = requests.get(url)

soup = BeautifulSoup(response.text, "html.parser")
elems = soup.find_all(href=re.compile("news.yahoo.co.jp/categories/sports/pickup"))
contents = soup.find_all('a', class_="sc-esjQYD")

#配列宣言(Webデータ挿入のため)
te = []
he = []

#Webのデータを挿入
for w in contents:
    te.append(w.get_text())
    he.append(w.get('href'))


# 現在日時の取得
dt_now = datetime.date.today() #- datetime.timedelta(days = 3)
print(dt_now)
#DBにデータ書き込み
for a,b in zip(te, he):
    cursor.execute("INSERT INTO rpa_data(title,url,now_date) VALUES(%s,%s,%s);",(a,b,dt_now))

# 保存を実行
connection.commit()

#辞書化する
cursor = connection.cursor(MySQLdb.cursors.DictCursor)

#データベースのデータを抽出
cursor.execute("SELECT * FROM rpa_data where now_date = '%s'" %dt_now)
result = cursor.fetchall()  #=> [{‘id’: 1, ‘name’: ‘foo’}, {‘id’: 2, ‘name’: ‘bar’}]


#配列宣言(DBのデータを挿入するため)
titles = []
ul = []
tmp_num = 1


#配列にDBに書き込んだデータを挿入
for d in result:
    titles.append(d.get('title'))
    ul.append(d.get('url'))
    tmp_num += 1


#エクセル生成
wbname = "goodlist.xlsx"
wb = excel.Workbook()
ws = wb.active
wb.create_sheet(title = str(dt_now)) #str()で文字列に変換
ws = wb[str(dt_now)] #文字列に変換した日付をシート名に格納


#エクセルシート情報
ws["A1"].value = 'タイトル'
ws["B1"].value = 'URL'

if ws == dt_now:
    #エクセルに書き込み
    for i in range(1, tmp_num):
        ws.cell(column=1, row=i+1, value=titles[i-1]) #配列titlesの要素をA列に出力
        ws.cell(column=2, row=i+1, value=ul[i-1]) #配列ulの要素をB列に出力


    wb.save("goodlist.xlsx")



if ws != dt_now:
    wbname = "goodlist.xlsx"
    wb = excel.load_workbook(wbname)
    ws = wb.active
    wb.create_sheet(title = str(dt_now)) #str()で文字列に変換
    ws = wb[str(dt_now)] #文字列に変換した日付をシート名に格納

    ws["A1"].value = 'タイトル'
    ws["B1"].value = 'URL'

    for i in range(1, tmp_num):
        ws.cell(column=1, row=i+1, value=titles[i-1]) #配列titlesの要素をA列に出力
        ws.cell(column=2, row=i+1, value=ul[i-1]) #配列ulの要素をB列に出力

    wb.save("goodlist.xlsx")

# データベース接続を閉じる
connection.close()