import MySQLdb
import openpyxl as excel
import datetime
from bs4.element import Comment
import requests
import glob
from bs4 import BeautifulSoup
import re
from selenium import webdriver




# データベースへの接続とカーソルの生成
connection = MySQLdb.connect(
    host='localhost',
    user='root',
    passwd='tkvip24week',
    db='pytest',
# テーブル内部で日本語を扱うために追加
    charset='utf8'
)


cursor = connection.cursor()

driver = webdriver.Chrome("C:/Users/原敬太/Documents/python_test/chromedriver")
url = 'https://news.yahoo.co.jp/categories/sports/'
driver.get(url)
#ドライバーを閉じる
driver.quit()

response = requests.get(url)

soup = BeautifulSoup(response.text, "html.parser")
elems = soup.find_all(href=re.compile("news.yahoo.co.jp/categories/sports/pickup"))
contents = soup.find_all('a', class_="sc-esjQYD")


#データベースのデータを抽出
cursor.execute("SELECT * FROM rpa_data")

#配列宣言
titles = []
ul = []
#niti = []

tmp_num = 1


# 現在日時の取得
dt_now = datetime.date.today()

#抽出したデータを配列に追加 (test[0]はDBのタイトル、test[1]はURLでDBの左から0.1.2....となっていく)
for w in contents:
    titles.append(w.get_text())
    ul.append(w.get('href'))
    tmp_num += 1



for a,b in zip(titles, ul):
    cursor.execute("INSERT INTO rpa_data(title,url) VALUES(%s,%s);",(a,b))

# 一覧の表示
#cursor.execute("SELECT * FROM rpa_data")

# 保存を実行
connection.commit()


# 現在日時の取得
dt_now = datetime.date.today()
print(dt_now)

#エクセル生成
wbname = "sports.xlsx"
wb = excel.Workbook()
ws = wb.active

#エクセルシート情報
ws["A1"] = dt_now
ws["A2"].value = 'タイトル'
ws["B2"].value = 'URL'
#ws["A1"].value = '日時'


#日時が変更された場合の条件文１
if ws["A1"] == dt_now:
    for i in range(1, tmp_num):
        ws.cell(column=1, row=i+1, value=titles[i-1]) #配列titlesの要素をA列に出力
        ws.cell(column=2, row=i+1, value=ul[i-1]) #配列ulの要素をB列に出力
#    ws.cell(column=3, row=i+1, value=niti[i-1]) #配列nitiの要素をC列に出力

#日時が変更された場合の条件文２
if ws["A1"] != dt_now:
    wbname = "sports.xlsx"
    wb = excel.load_workbook(wbname) #指定したエクセルの読み込み
    ws = wb.active
    wb.create_sheet()
    ws = wb['Sheet1']

    #エクセルシート情報
    ws["A1"] = dt_now
    ws["A2"].value = 'タイトル'
    ws["B2"].value = 'URL'

    for i in range(1, tmp_num):
        ws.cell(column=1, row=i+1, value=titles[i-1]) #配列titlesの要素をA列に出力
        ws.cell(column=2, row=i+1, value=ul[i-1]) #配列ulの要素をB列に出力
#    ws.cell(column=3, row=i+1, value=niti[i-1]) #配列nitiの要素をC列に出力



wb.save("sports.xlsx")

# データベース接続を閉じる
connection.close()