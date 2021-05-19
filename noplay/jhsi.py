from bs4.element import Comment
import requests
import glob
from bs4 import BeautifulSoup
import re
from selenium import webdriver
import MySQLdb

# データベースへの接続とカーソルの生成
connection = MySQLdb.connect(
    host='localhost',
    user='root',
    passwd='tkvip24week',
    db='pytest',
# テーブル内部で日本語を扱うために追加
    charset='utf8'
)

cursor = connection.cursor()

driver = webdriver.Chrome("C:/Users/原敬太/Documents/python_test/chromedriver")
url = 'https://news.yahoo.co.jp/categories/sports/'
driver.get(url)
#ドライバーを閉じる
driver.quit()

response = requests.get(url)

soup = BeautifulSoup(response.text, "html.parser")
elems = soup.find_all(href=re.compile("news.yahoo.co.jp/categories/sports/pickup"))
contents = soup.find_all('a', class_="sc-esjQYD")

titles = []
ul = []

tmp_num = 1

for w in contents:
    titles.append(w.get_text())
    ul.append(w.get('href'))
    tmp_num += 1


for a,b in zip(titles, ul):
    cursor.execute("INSERT INTO rpa_data(title,url) VALUES(%s,%s);",(a,b))

# 一覧の表示
cursor.execute("SELECT * FROM rpa_data")

for row in cursor:
    print(row)

# 保存を実行
connection.commit()

#エクセルに書き込み
import openpyxl as excel

wbname = "sports.xlsx"
wb = excel.Workbook()
ws = wb.active

ws["A1"].value = 'タイトル'
ws["B1"].value = 'URL'


for i in range(1, tmp_num):
    ws.cell(column=1, row=i+1, value=titles[i-1]) #配列titlesの要素をA列に出力
    ws.cell(column=2, row=i+1, value=ul[i-1]) #配列ulの要素をB列に出力

# データベース接続を閉じる
connection.close()
wb.save("sports.xlsx")

