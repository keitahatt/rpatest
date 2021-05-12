from bs4.element import Comment
import requests
import glob
from bs4 import BeautifulSoup
import re
from selenium import webdriver

driver = webdriver.Chrome("C:/Users/原敬太/Documents/python_test/chromedriver")
url = 'https://news.yahoo.co.jp/categories/sports/'
driver.get(url)
#ドライバーを閉じる
driver.quit()

response = requests.get(url)

soup = BeautifulSoup(response.text, "html.parser")
elems = soup.find_all(href=re.compile("news.yahoo.co.jp/categories/sports/pickup"))
contents = soup.find_all('a', class_="sc-esjQYD")

titles = []
ul = []

tmp_num = 1

for w in contents:
    titles.append(w.get_text())
    ul.append(w.get('href')) 
    tmp_num += 1

import openpyxl as excel

wbname = "sports.xlsx"
wb = excel.Workbook()
ws = wb.active

ws["A1"].value = 'タイトル'
ws["B1"].value = 'URL'


for i in range(1, tmp_num):
    ws.cell(column=1, row=i+1, value=titles[i-1]) #配列titlesの要素をA列に出力
    ws.cell(column=2, row=i+1, value=ul[i-1]) #配列ulの要素をB列に出力

wb.save("sports.xlsx")
